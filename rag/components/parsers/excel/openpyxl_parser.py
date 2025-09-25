"""Excel parser using OpenPyXL library."""

from pathlib import Path
from typing import Dict, Any, List, Optional
import logging

logger = logging.getLogger(__name__)


class ExcelParser_OpenPyXL:
    """Excel parser using OpenPyXL for XLSX files."""

    def __init__(
        self,
        name: str = "ExcelParser_OpenPyXL",
        config: Optional[Dict[str, Any]] = None,
    ):
        self.name = name
        self.config = config or {}
        self.chunk_size = self.config.get("chunk_size", 1000)
        self.extract_formulas = self.config.get("extract_formulas", False)
        self.extract_metadata = self.config.get("extract_metadata", True)
        self.sheets = self.config.get("sheets", None)  # None = all sheets
        self.data_only = self.config.get(
            "data_only", True
        )  # True = values, False = formulas

    def validate_config(self) -> bool:
        """Validate configuration."""
        return True

    def parse(self, source: str, **kwargs):
        """Parse Excel using OpenPyXL."""
        from core.base import Document, ProcessingResult

        try:
            import openpyxl
        except ImportError:
            return ProcessingResult(
                documents=[],
                errors=[
                    {
                        "error": "OpenPyXL not installed. Install with: pip install openpyxl",
                        "source": source,
                    }
                ],
            )

        path = Path(source)
        if not path.exists():
            return ProcessingResult(
                documents=[],
                errors=[{"error": f"File not found: {source}", "source": source}],
            )

        try:
            # Load workbook
            wb = openpyxl.load_workbook(source, data_only=self.data_only)

            # Get sheets to process
            if self.sheets:
                sheet_names = [s for s in self.sheets if s in wb.sheetnames]
            else:
                sheet_names = wb.sheetnames

            documents = []

            for sheet_name in sheet_names:
                ws = wb[sheet_name]

                # Extract sheet data
                rows = []
                for row in ws.iter_rows(values_only=self.data_only):
                    # Filter out empty rows
                    if any(cell is not None for cell in row):
                        rows.append(
                            [str(cell) if cell is not None else "" for cell in row]
                        )

                if not rows:
                    continue

                metadata = {
                    "source": str(path),
                    "file_name": path.name,
                    "parser": self.name,
                    "tool": "OpenPyXL",
                    "sheet_name": sheet_name,
                    "total_sheets": len(wb.sheetnames),
                    "rows": len(rows),
                    "columns": len(rows[0]) if rows else 0,
                }

                if self.extract_metadata:
                    # Add workbook properties
                    props = wb.properties
                    if props:
                        metadata.update(
                            {
                                "title": props.title,
                                "creator": props.creator,
                                "created": str(props.created)
                                if props.created
                                else None,
                                "modified": str(props.modified)
                                if props.modified
                                else None,
                            }
                        )

                    # Add sheet-specific metadata
                    metadata.update(
                        {
                            "max_row": ws.max_row,
                            "max_column": ws.max_column,
                            "merged_cells": len(ws.merged_cells.ranges),
                        }
                    )

                # Format data as text
                text = self._format_sheet(rows, sheet_name)

                # Apply chunking if needed
                if (
                    self.chunk_size
                    and self.chunk_size > 0
                    and len(rows) > self.chunk_size
                ):
                    for chunk_idx in range(0, len(rows), self.chunk_size):
                        chunk_rows = rows[chunk_idx : chunk_idx + self.chunk_size]
                        chunk_text = self._format_sheet(chunk_rows, sheet_name)

                        chunk_metadata = metadata.copy()
                        chunk_metadata.update(
                            {
                                "chunk_index": chunk_idx // self.chunk_size,
                                "chunk_rows": len(chunk_rows),
                                "start_row": chunk_idx,
                                "end_row": min(chunk_idx + self.chunk_size, len(rows)),
                            }
                        )

                        doc = Document(
                            content=chunk_text,
                            metadata=chunk_metadata,
                            id=f"{path.stem}_{sheet_name}_chunk_{chunk_idx // self.chunk_size + 1}",
                            source=str(path),
                        )
                        documents.append(doc)
                else:
                    doc = Document(
                        content=text,
                        metadata=metadata,
                        id=f"{path.stem}_{sheet_name}",
                        source=str(path),
                    )
                    documents.append(doc)

            wb.close()

            return ProcessingResult(
                documents=documents,
                errors=[],
                metrics={
                    "total_documents": len(documents),
                    "parser_type": self.name,
                    "tool": "OpenPyXL",
                    "sheets_processed": len(sheet_names),
                },
            )

        except Exception as e:
            logger.error(f"Failed to parse {source}: {e}")
            return ProcessingResult(
                documents=[], errors=[{"error": str(e), "source": source}]
            )

    def _format_sheet(self, rows: List[List[str]], sheet_name: str) -> str:
        """Format sheet data as text table."""
        if not rows:
            return f"Sheet: {sheet_name}\n(empty)"

        lines = [f"Sheet: {sheet_name}", "=" * 50]

        # Calculate column widths
        col_widths = []
        num_cols = max(len(row) for row in rows)

        for col_idx in range(num_cols):
            max_width = 0
            for row in rows:
                if col_idx < len(row):
                    max_width = max(max_width, len(str(row[col_idx])))
            col_widths.append(min(max_width, 30))  # Cap at 30 chars

        # Format rows
        for row_idx, row in enumerate(rows):
            row_line = " | ".join(
                str(row[i] if i < len(row) else "").ljust(col_widths[i])[
                    : col_widths[i]
                ]
                for i in range(num_cols)
            )
            lines.append(row_line)

            # Add separator after header row
            if row_idx == 0:
                lines.append("-" * len(row_line))

        return "\n".join(lines)
